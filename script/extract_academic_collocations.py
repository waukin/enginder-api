from openpyxl import load_workbook
import csv
import re

INPUT_FILE = "The_Academic_Collocation_List.xlsx"
SHEET_NAME = "Academic Collocation List"
OUTPUT_FILE = "academic_collocations.csv"

POS_TAGS = {"adj", "v", "vpp", "adv", "n"}
POS_PATTERN = re.compile(r"\(([^)]+)\)")


def clean_token(text):
    if not text:
        return None

    text = str(text)

    def replacer(match):
        tag = match.group(1).strip().lower()
        return "" if tag in POS_TAGS else match.group(0)

    return POS_PATTERN.sub(replacer, text).strip()


def extract_collocations():
    wb = load_workbook(INPUT_FILE, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]

    current_left = None
    rows = []

    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx < 2:  # skip first two rows
            continue

        left_raw = row[2] if len(row) > 2 else None  # Column C
        right_raw = row[3] if len(row) > 3 else None  # Column D

        left = clean_token(left_raw)
        right = clean_token(right_raw)

        if left:
            current_left = left

        if current_left and right:
            rows.append((current_left, right))

    with open(OUTPUT_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        # writer.writerow(["left", "right"])
        writer.writerows(rows)

    print(f"Extracted {len(rows)} collocations to {OUTPUT_FILE}")


if __name__ == "__main__":
    extract_collocations()
